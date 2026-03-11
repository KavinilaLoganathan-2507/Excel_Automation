"""
Report Generator Module - Auto-generate comprehensive reports
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional
from datetime import datetime
import io
import base64

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True  
except ImportError:
    HAS_OPENPYXL = False


class ReportGenerator:
    """Generate comprehensive reports in multiple formats"""
    
    def __init__(self, df: pd.DataFrame, eda_results: Dict = None, analysis: str = None):
        self.df = df
        self.eda_results = eda_results or {}
        self.analysis = analysis or ""
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def generate_html_report(self) -> str:
        """Generate a comprehensive HTML report"""
        html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>TechXcel Data Report</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 2rem;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 20px;
            box-shadow: 0 25px 50px rgba(0,0,0,0.15);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #8b5cf6, #06b6d4);
            color: white;
            padding: 3rem;
            text-align: center;
        }}
        .header h1 {{
            font-size: 2.5rem;
            margin-bottom: 0.5rem;
        }}
        .header p {{
            opacity: 0.9;
        }}
        .content {{
            padding: 2rem;
        }}
        .section {{
            margin-bottom: 2rem;
            padding: 1.5rem;
            background: #f8fafc;
            border-radius: 12px;
            border-left: 4px solid #8b5cf6;
        }}
        .section h2 {{
            color: #1e293b;
            margin-bottom: 1rem;
            font-size: 1.5rem;
        }}
        .metrics-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1.5rem;
            margin-top: 1rem;
        }}
        .metric-card {{
            background: white;
            padding: 1.5rem;
            border-radius: 12px;
            text-align: center;
            box-shadow: 0 4px 6px rgba(0,0,0,0.07);
        }}
        .metric-value {{
            font-size: 2rem;
            font-weight: 700;
            color: #8b5cf6;
        }}
        .metric-label {{
            color: #64748b;
            font-size: 0.875rem;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 1rem;
        }}
        th, td {{
            padding: 0.75rem 1rem;
            text-align: left;
            border-bottom: 1px solid #e2e8f0;
        }}
        th {{
            background: #8b5cf6;
            color: white;
            font-weight: 600;
        }}
        tr:hover {{
            background: #f1f5f9;
        }}
        .insight {{
            padding: 1rem;
            margin: 0.5rem 0;
            border-radius: 8px;
            display: flex;
            align-items: flex-start;
            gap: 0.75rem;
        }}
        .insight.success {{ background: #dcfce7; border-left: 4px solid #22c55e; }}
        .insight.warning {{ background: #fef3c7; border-left: 4px solid #f59e0b; }}
        .insight.danger {{ background: #fee2e2; border-left: 4px solid #ef4444; }}
        .insight.info {{ background: #dbeafe; border-left: 4px solid #3b82f6; }}
        .insight-icon {{ font-size: 1.25rem; }}
        .insight-content h4 {{ color: #1e293b; margin-bottom: 0.25rem; }}
        .insight-content p {{ color: #64748b; font-size: 0.875rem; }}
        .footer {{
            background: #1e293b;
            color: white;
            padding: 2rem;
            text-align: center;
        }}
        .quality-score {{
            width: 120px;
            height: 120px;
            border-radius: 50%;
            background: conic-gradient(#8b5cf6 var(--score), #e2e8f0 0);
            display: flex;
            align-items: center;
            justify-content: center;
            margin: 1rem auto;
        }}
        .quality-score-inner {{
            width: 100px;
            height: 100px;
            border-radius: 50%;
            background: white;
            display: flex;
            align-items: center;
            justify-content: center;
            flex-direction: column;
        }}
        .quality-score-value {{ font-size: 1.75rem; font-weight: 700; color: #8b5cf6; }}
        .quality-score-label {{ font-size: 0.75rem; color: #64748b; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>⚡ TechXcel Data Report</h1>
            <p>Generated: {self.timestamp}</p>
        </div>
        
        <div class="content">
            {self._generate_overview_section()}
            {self._generate_quality_section()}
            {self._generate_statistics_section()}
            {self._generate_correlations_section()}
            {self._generate_insights_section()}
            {self._generate_data_preview_section()}
        </div>
        
        <div class="footer">
            <p>Generated by TechXcel - Smart Excel Analytics with AI</p>
        </div>
    </div>
</body>
</html>
        """
        return html
    
    def _generate_overview_section(self) -> str:
        """Generate overview section HTML"""
        overview = self.eda_results.get('overview', {})
        shape = overview.get('shape', {})
        column_types = overview.get('column_types', {})
        
        return f"""
        <div class="section">
            <h2>📊 Data Overview</h2>
            <div class="metrics-grid">
                <div class="metric-card">
                    <div class="metric-value">{shape.get('rows', len(self.df))}</div>
                    <div class="metric-label">Total Rows</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{shape.get('columns', len(self.df.columns))}</div>
                    <div class="metric-label">Total Columns</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{column_types.get('numeric', 0)}</div>
                    <div class="metric-label">Numeric Columns</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{column_types.get('categorical', 0)}</div>
                    <div class="metric-label">Categorical Columns</div>
                </div>
            </div>
        </div>
        """
    
    def _generate_quality_section(self) -> str:
        """Generate data quality section HTML"""
        quality = self.eda_results.get('data_quality', {})
        score = quality.get('overall_score', 100)
        
        return f"""
        <div class="section">
            <h2>🎯 Data Quality</h2>
            <div style="display: flex; align-items: center; gap: 3rem; flex-wrap: wrap;">
                <div class="quality-score" style="--score: {score * 3.6}deg;">
                    <div class="quality-score-inner">
                        <div class="quality-score-value">{score}%</div>
                        <div class="quality-score-label">Grade: {quality.get('grade', 'N/A')}</div>
                    </div>
                </div>
                <div class="metrics-grid" style="flex: 1;">
                    <div class="metric-card">
                        <div class="metric-value">{quality.get('completeness', 100)}%</div>
                        <div class="metric-label">Completeness</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{quality.get('uniqueness', 100)}%</div>
                        <div class="metric-label">Uniqueness</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{quality.get('consistency', 100)}%</div>
                        <div class="metric-label">Consistency</div>
                    </div>
                </div>
            </div>
        </div>
        """
    
    def _generate_statistics_section(self) -> str:
        """Generate statistics section HTML"""
        numeric_summary = self.eda_results.get('numeric_summary', {})
        
        if isinstance(numeric_summary, dict) and 'message' not in numeric_summary:
            rows = ""
            for col, stats in numeric_summary.items():
                rows += f"""
                <tr>
                    <td><strong>{col}</strong></td>
                    <td>{stats.get('mean', 'N/A')}</td>
                    <td>{stats.get('median', 'N/A')}</td>
                    <td>{stats.get('std', 'N/A')}</td>
                    <td>{stats.get('min', 'N/A')}</td>
                    <td>{stats.get('max', 'N/A')}</td>
                </tr>
                """
            
            return f"""
            <div class="section">
                <h2>📈 Statistical Summary</h2>
                <table>
                    <thead>
                        <tr>
                            <th>Column</th>
                            <th>Mean</th>
                            <th>Median</th>
                            <th>Std Dev</th>
                            <th>Min</th>
                            <th>Max</th>
                        </tr>
                    </thead>
                    <tbody>
                        {rows}
                    </tbody>
                </table>
            </div>
            """
        return ""
    
    def _generate_correlations_section(self) -> str:
        """Generate correlations section HTML"""
        correlations = self.eda_results.get('correlations', {})
        strong_corrs = correlations.get('strong_correlations', [])
        
        if strong_corrs:
            rows = ""
            for corr in strong_corrs[:10]:
                color = "#22c55e" if corr['direction'] == 'positive' else "#ef4444"
                rows += f"""
                <tr>
                    <td>{corr['var1']}</td>
                    <td>{corr['var2']}</td>
                    <td style="color: {color}; font-weight: 600;">{corr['correlation']}</td>
                    <td>{corr['strength'].capitalize()}</td>
                </tr>
                """
            
            return f"""
            <div class="section">
                <h2>🔗 Key Correlations</h2>
                <table>
                    <thead>
                        <tr>
                            <th>Variable 1</th>
                            <th>Variable 2</th>
                            <th>Correlation</th>
                            <th>Strength</th>
                        </tr>
                    </thead>
                    <tbody>
                        {rows}
                    </tbody>
                </table>
            </div>
            """
        return ""
    
    def _generate_insights_section(self) -> str:
        """Generate insights section HTML"""
        # Generate insights from EDA if not provided
        insights_html = ""
        
        # Add AI analysis if available
        if self.analysis:
            insights_html += f"""
            <div class="insight info">
                <div class="insight-icon">🤖</div>
                <div class="insight-content">
                    <h4>AI Analysis</h4>
                    <p>{self.analysis[:500]}...</p>
                </div>
            </div>
            """
        
        return f"""
        <div class="section">
            <h2>💡 Key Insights</h2>
            {insights_html}
        </div>
        """ if insights_html else ""
    
    def _generate_data_preview_section(self) -> str:
        """Generate data preview section HTML"""
        preview = self.df.head(10)
        
        headers = "".join([f"<th>{col}</th>" for col in preview.columns])
        rows = ""
        for _, row in preview.iterrows():
            cells = "".join([f"<td>{val}</td>" for val in row])
            rows += f"<tr>{cells}</tr>"
        
        return f"""
        <div class="section">
            <h2>📋 Data Preview (First 10 Rows)</h2>
            <div style="overflow-x: auto;">
                <table>
                    <thead>
                        <tr>{headers}</tr>
                    </thead>
                    <tbody>
                        {rows}
                    </tbody>
                </table>
            </div>
        </div>
        """
    
    def generate_excel_dashboard(self) -> bytes:
        """Generate an Excel dashboard with charts"""
        if not HAS_OPENPYXL:
            return None
        
        output = io.BytesIO()
        wb = Workbook()
        
        # Dashboard sheet
        ws_dashboard = wb.active
        ws_dashboard.title = "Dashboard"
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        metric_font = Font(bold=True, size=14)
        
        # Title
        ws_dashboard["A1"] = "TechXcel Dashboard"
        ws_dashboard["A1"].font = Font(bold=True, size=20, color="8B5CF6")
        ws_dashboard.merge_cells("A1:F1")
        
        ws_dashboard["A2"] = f"Generated: {self.timestamp}"
        ws_dashboard["A2"].font = Font(italic=True, color="666666")
        
        # Overview metrics
        row = 4
        ws_dashboard[f"A{row}"] = "Data Overview"
        ws_dashboard[f"A{row}"].font = Font(bold=True, size=14)
        
        row += 1
        metrics = [
            ("Total Rows", len(self.df)),
            ("Total Columns", len(self.df.columns)),
            ("Numeric Columns", len(self.df.select_dtypes(include=[np.number]).columns)),
            ("Memory Usage", f"{self.df.memory_usage(deep=True).sum() / 1024:.1f} KB")
        ]
        
        for i, (label, value) in enumerate(metrics):
            col = chr(65 + i * 2)  # A, C, E, G
            ws_dashboard[f"{col}{row}"] = label
            ws_dashboard[f"{col}{row}"].font = header_font
            ws_dashboard[f"{col}{row}"].fill = header_fill
            ws_dashboard[f"{col}{row+1}"] = value
            ws_dashboard[f"{col}{row+1}"].font = metric_font
        
        # Data sheet
        ws_data = wb.create_sheet("Data")
        for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Statistics sheet
        ws_stats = wb.create_sheet("Statistics")
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        
        if len(numeric_cols) > 0:
            stats_df = self.df[numeric_cols].describe()
            
            for r_idx, row in enumerate(dataframe_to_rows(stats_df, index=True, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_stats.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
            
            # Add a bar chart for the first numeric column means
            if len(numeric_cols) > 1:
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Column Means Comparison"
                
                # Find the mean row
                mean_row = 3  # Usually row 3 in describe()
                data = Reference(ws_stats, min_col=2, min_row=mean_row, max_col=len(numeric_cols)+1, max_row=mean_row)
                cats = Reference(ws_stats, min_col=2, min_row=1, max_col=len(numeric_cols)+1, max_row=1)
                
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(cats)
                chart.shape = 4
                chart.width = 15
                chart.height = 10
                
                ws_stats.add_chart(chart, "A15")
        
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def generate_csv_report(self) -> str:
        """Generate CSV export"""
        return self.df.to_csv(index=False)
    
    def generate_markdown_report(self) -> str:
        """Generate markdown report"""
        md = f"""# TechXcel Data Report

**Generated:** {self.timestamp}

## 📊 Data Overview

| Metric | Value |
|--------|-------|
| Total Rows | {len(self.df)} |
| Total Columns | {len(self.df.columns)} |
| Memory Usage | {self.df.memory_usage(deep=True).sum() / 1024:.2f} KB |

## 📈 Statistical Summary

"""
        # Add statistics table
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            stats = self.df[numeric_cols].describe()
            md += stats.to_markdown()
        
        md += f"""

## 📋 Column Information

| Column | Type |
|--------|------|
"""
        for col in self.df.columns:
            md += f"| {col} | {self.df[col].dtype} |\n"
        
        if self.analysis:
            md += f"""

## 🤖 AI Analysis

{self.analysis}
"""
        
        return md
    
    def get_report_download_data(self, format: str = 'html') -> tuple:
        """Get report data for download"""
        if format == 'html':
            content = self.generate_html_report()
            return content, 'text/html', 'report.html'
        elif format == 'excel':
            content = self.generate_excel_dashboard()
            return content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'dashboard.xlsx'
        elif format == 'csv':
            content = self.generate_csv_report()
            return content, 'text/csv', 'data.csv'
        elif format == 'markdown':
            content = self.generate_markdown_report()
            return content, 'text/markdown', 'report.md'
        else:
            return None, None, None


class DashboardBuilder:
    """Build interactive dashboard views"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df
    
    def get_kpi_cards(self) -> List[Dict[str, Any]]:
        """Generate KPI cards data"""
        numeric_cols = self.df.select_dtypes(include=[np.number]).columns
        
        kpis = []
        for col in numeric_cols:
            data = self.df[col].dropna()
            if len(data) > 1:
                # Calculate trend
                first_half = data.iloc[:len(data)//2].mean()
                second_half = data.iloc[len(data)//2:].mean()
                trend = ((second_half - first_half) / first_half * 100) if first_half != 0 else 0
                
                kpis.append({
                    "name": col,
                    "value": round(float(data.sum()), 2),
                    "average": round(float(data.mean()), 2),
                    "trend": round(float(trend), 1),
                    "trend_direction": "up" if trend > 0 else "down" if trend < 0 else "neutral"
                })
        
        return kpis
    
    def get_chart_recommendations(self) -> List[Dict[str, Any]]:
        """Recommend appropriate chart types for the data"""
        recommendations = []
        
        numeric_cols = list(self.df.select_dtypes(include=[np.number]).columns)
        categorical_cols = list(self.df.select_dtypes(include=['object', 'category']).columns)
        
        # Time series recommendation
        date_cols = [col for col in self.df.columns if pd.api.types.is_datetime64_any_dtype(self.df[col])]
        if date_cols and numeric_cols:
            recommendations.append({
                "chart_type": "line",
                "title": f"{numeric_cols[0]} Over Time",
                "x": date_cols[0],
                "y": numeric_cols[0],
                "description": "Track trends over time"
            })
        
        # Correlation scatter
        if len(numeric_cols) >= 2:
            recommendations.append({
                "chart_type": "scatter",
                "title": f"{numeric_cols[0]} vs {numeric_cols[1]}",
                "x": numeric_cols[0],
                "y": numeric_cols[1],
                "description": "Explore relationships between variables"
            })
        
        # Category distribution
        if categorical_cols and numeric_cols:
            recommendations.append({
                "chart_type": "bar",
                "title": f"{numeric_cols[0]} by {categorical_cols[0]}",
                "x": categorical_cols[0],
                "y": numeric_cols[0],
                "description": "Compare values across categories"
            })
        
        # Distribution histogram
        if numeric_cols:
            recommendations.append({
                "chart_type": "histogram",
                "title": f"Distribution of {numeric_cols[0]}",
                "column": numeric_cols[0],
                "description": "Understand data distribution"
            })
        
        # Pie chart for categorical
        if categorical_cols:
            recommendations.append({
                "chart_type": "pie",
                "title": f"{categorical_cols[0]} Distribution",
                "column": categorical_cols[0],
                "description": "Show category proportions"
            })
        
        return recommendations
